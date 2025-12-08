"""
Tender Management Module with AI Enhancement
=============================================

Provides:
- Load and parse OCR results from JSON/Excel
- AI-powered data cleaning (Gemini + Groq free tier)
- Dashboard statistics and analytics
- Search and filtering
- Export functionality
"""

import os
import sys
import json
import re
import time
import logging
import threading
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Any, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict, field

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION
# ============================================================================

CONFIG = {
    "ocr_results_dir": os.path.expanduser("~/Documents/MOH Tenders/ocr_results"),
    "ai_enhanced_dir": os.path.expanduser("~/Documents/MOH Tenders/ai_enhanced"),
    "cache_file": os.path.expanduser("~/Documents/MOH Tenders/tender_cache.json"),
}

# API Keys - load from environment variables
# Set these in your environment or .env file:
#   export GROQ_API_KEY=your_key_here
#   export GEMINI_API_KEY=your_key_here
API_KEYS = {
    "GROQ_API_KEY": os.environ.get("GROQ_API_KEY", ""),
    "GEMINI_API_KEY": os.environ.get("GEMINI_API_KEY", ""),
}

# ============================================================================
# AI PROVIDER IMPORTS
# ============================================================================

GEMINI_AVAILABLE = False
GROQ_AVAILABLE = False

try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    pass

try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    pass

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class TenderItem:
    item_number: str
    description: str
    quantity: str
    unit: str
    ai_cleaned: bool = False
    original_description: str = ""
    confidence: float = 1.0


@dataclass
class Tender:
    reference: str
    department: str
    title: str
    closing_date: str
    items: List[TenderItem]
    source_file: str
    ocr_confidence: float
    has_arabic: bool = False
    ai_enhanced: bool = False
    ai_provider: str = ""
    extraction_date: str = ""


@dataclass
class DashboardStats:
    total_tenders: int
    total_items: int
    unique_departments: int
    ai_enhanced_count: int
    avg_confidence: float
    items_per_tender: float
    department_breakdown: Dict[str, int]
    recent_tenders: List[Dict]
    top_items: List[Dict]


# ============================================================================
# AI ENHANCEMENT ENGINE
# ============================================================================

class AIWorker:
    """Base AI worker with rate limiting"""

    def __init__(self, name: str, rpm_limit: int):
        self.name = name
        self.rpm_limit = rpm_limit
        self.min_interval = 60.0 / rpm_limit
        self.last_request = 0
        self.lock = threading.Lock()
        self.request_count = 0
        self.error_count = 0

    def _wait_for_rate_limit(self):
        with self.lock:
            elapsed = time.time() - self.last_request
            if elapsed < self.min_interval:
                time.sleep(self.min_interval - elapsed)
            self.last_request = time.time()
            self.request_count += 1

    def call(self, prompt: str) -> Optional[str]:
        raise NotImplementedError


class GeminiWorker(AIWorker):
    def __init__(self, api_key: str):
        super().__init__("gemini", rpm_limit=15)
        self.model = None
        if GEMINI_AVAILABLE and api_key:
            try:
                genai.configure(api_key=api_key)
                self.model = genai.GenerativeModel("gemini-2.0-flash")
            except Exception as e:
                logger.error(f"Gemini init failed: {e}")

    def call(self, prompt: str) -> Optional[str]:
        if not self.model:
            return None
        self._wait_for_rate_limit()
        try:
            response = self.model.generate_content(prompt)
            return response.text.strip()
        except Exception as e:
            self.error_count += 1
            return None


class GroqWorker(AIWorker):
    def __init__(self, api_key: str):
        super().__init__("groq", rpm_limit=30)
        self.client = None
        if GROQ_AVAILABLE and api_key:
            try:
                self.client = Groq(api_key=api_key)
            except Exception as e:
                logger.error(f"Groq init failed: {e}")

    def call(self, prompt: str) -> Optional[str]:
        if not self.client:
            return None
        self._wait_for_rate_limit()
        try:
            response = self.client.chat.completions.create(
                model="llama-3.1-8b-instant",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=4000
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            self.error_count += 1
            return None


class AIEngine:
    """AI Enhancement Engine with round-robin load balancing"""

    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        if self._initialized:
            return
        self._initialized = True

        self.workers = []
        self.worker_index = 0
        self.worker_lock = threading.Lock()

        # Initialize free tier workers
        gemini_key = API_KEYS.get("GEMINI_API_KEY", "")
        groq_key = API_KEYS.get("GROQ_API_KEY", "")

        if gemini_key and GEMINI_AVAILABLE:
            worker = GeminiWorker(gemini_key)
            if worker.model:
                self.workers.append(worker)
                logger.info("Gemini worker ready (15 RPM)")

        if groq_key and GROQ_AVAILABLE:
            worker = GroqWorker(groq_key)
            if worker.client:
                self.workers.append(worker)
                logger.info("Groq worker ready (30 RPM)")

        self.total_rpm = sum(w.rpm_limit for w in self.workers)
        logger.info(f"AI Engine: {len(self.workers)} workers, {self.total_rpm} RPM")

    def _get_worker(self) -> Optional[AIWorker]:
        if not self.workers:
            return None
        with self.worker_lock:
            worker = self.workers[self.worker_index % len(self.workers)]
            self.worker_index += 1
            return worker

    def clean_description(self, description: str) -> Tuple[str, bool]:
        """Clean OCR item description using AI"""
        if not description or len(description) < 5:
            return description, False

        worker = self._get_worker()
        if not worker:
            return description, False

        prompt = f"""Clean this medical tender item description from OCR noise.

Original: "{description}"

Rules:
1. Fix OCR errors (0/O, 1/l/I, rn/m)
2. Remove garbage text, page numbers, dimensions that aren't part of item
3. Keep medical terms accurate (LATEX, STERILE, etc.)
4. Keep size info if relevant (5ML, MEDIUM)
5. Remove document artifacts

Return ONLY the cleaned description. If garbage, return "INVALID"."""

        result = worker.call(prompt)
        if result:
            result = result.strip().strip('"\'')
            if result != "INVALID" and len(result) >= 5:
                return result, result.lower() != description.lower()

        return description, False

    def extract_items(self, ocr_text: str) -> List[Dict]:
        """Extract items from OCR text using AI"""
        worker = self._get_worker()
        if not worker:
            return []

        prompt = f"""Extract medical tender items from this Kuwait MOH document.

Each item needs: item_number, description, quantity, unit

OCR Text:
---
{ocr_text[:5000]}
---

Return JSON array:
[{{"item_number": "1", "description": "SURGICAL GLOVES LATEX", "quantity": "5000", "unit": "PCS"}}]

Return ONLY JSON array. If no items, return []"""

        result = worker.call(prompt)
        if result:
            try:
                match = re.search(r'\[[\s\S]*\]', result)
                if match:
                    items = json.loads(match.group())
                    return [i for i in items if all(k in i for k in ['item_number', 'description', 'quantity', 'unit'])]
            except:
                pass
        return []

    def get_stats(self) -> Dict:
        return {
            "workers": [{"name": w.name, "rpm": w.rpm_limit, "requests": w.request_count, "errors": w.error_count} for w in self.workers],
            "total_rpm": self.total_rpm,
            "available": len(self.workers) > 0
        }


# ============================================================================
# TENDER MANAGER
# ============================================================================

class TenderManager:
    """
    Central manager for all tender operations.

    Features:
    - Load OCR results from files
    - Cache for fast access
    - AI enhancement on demand
    - Search and filtering
    - Dashboard analytics
    """

    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        if self._initialized:
            return
        self._initialized = True

        self.tenders: Dict[str, Dict] = {}  # reference -> tender data
        self.ai_engine = AIEngine()
        self.cache_file = Path(CONFIG["cache_file"])
        self.ocr_dir = Path(CONFIG["ocr_results_dir"])

        # Load cache or OCR results
        self._load_data()

    def _load_data(self):
        """Load tenders from cache or OCR results"""
        # Try cache first
        if self.cache_file.exists():
            try:
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.tenders = data.get('tenders', {})
                    logger.info(f"Loaded {len(self.tenders)} tenders from cache")
                    return
            except:
                pass

        # Load from OCR results
        self.reload_from_ocr()

    def reload_from_ocr(self):
        """Reload all tenders from OCR result files"""
        self.tenders = {}

        if not self.ocr_dir.exists():
            logger.warning(f"OCR results directory not found: {self.ocr_dir}")
            return

        # Load combined file first if exists
        combined_files = list(self.ocr_dir.glob("ocr_combined_*.json"))
        if combined_files:
            latest = max(combined_files, key=lambda p: p.stat().st_mtime)
            self._load_json_file(latest)
        else:
            # Load batch files
            for json_file in sorted(self.ocr_dir.glob("ocr_batch_*.json")):
                self._load_json_file(json_file)

        # Also try Excel files
        for xlsx_file in sorted(self.ocr_dir.glob("*.xlsx")):
            self._load_excel_file(xlsx_file)

        logger.info(f"Loaded {len(self.tenders)} unique tenders from OCR results")
        self._save_cache()

    def _load_json_file(self, filepath: Path):
        """Load tenders from JSON file"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if isinstance(data, list):
                for tender in data:
                    ref = tender.get('reference', tender.get('Tender Reference', ''))
                    if ref and ref != 'UNKNOWN':
                        self.tenders[ref] = self._normalize_tender(tender)
            elif isinstance(data, dict):
                for tender in data.get('tenders', []):
                    ref = tender.get('reference', '')
                    if ref and ref != 'UNKNOWN':
                        self.tenders[ref] = self._normalize_tender(tender)

        except Exception as e:
            logger.error(f"Error loading {filepath}: {e}")

    def _load_excel_file(self, filepath: Path):
        """Load tenders from Excel file"""
        if not OPENPYXL_AVAILABLE:
            return

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            tenders_dict = {}

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                ref = row_dict.get('Tender Reference', '')

                if not ref or ref == 'UNKNOWN':
                    continue

                if ref not in tenders_dict:
                    tenders_dict[ref] = {
                        'reference': ref,
                        'department': row_dict.get('Department', ''),
                        'title': row_dict.get('Title', ''),
                        'closing_date': str(row_dict.get('Closing Date', '') or ''),
                        'source_file': row_dict.get('Source File', ''),
                        'ocr_confidence': float(row_dict.get('OCR Confidence', 0) or 0),
                        'has_arabic': bool(row_dict.get('Has Arabic', False)),
                        'items': [],
                        'ai_enhanced': False,
                        'extraction_date': str(row_dict.get('Extraction Date', '') or '')
                    }

                # Add item
                desc = row_dict.get('Item Description', '')
                if desc and len(str(desc)) > 3:
                    tenders_dict[ref]['items'].append({
                        'item_number': str(row_dict.get('Item #', '') or ''),
                        'description': str(desc),
                        'quantity': str(row_dict.get('Quantity', '') or ''),
                        'unit': str(row_dict.get('Unit', 'PCS') or 'PCS'),
                        'ai_cleaned': False
                    })

            # Merge into main dict
            for ref, tender in tenders_dict.items():
                if ref not in self.tenders:
                    self.tenders[ref] = tender

            wb.close()

        except Exception as e:
            logger.error(f"Error loading Excel {filepath}: {e}")

    def _normalize_tender(self, tender: Dict) -> Dict:
        """Normalize tender data structure"""
        return {
            'reference': tender.get('reference', tender.get('Tender Reference', '')),
            'department': tender.get('department', tender.get('Department', '')),
            'title': tender.get('title', tender.get('Title', '')),
            'closing_date': str(tender.get('closing_date', tender.get('Closing Date', '')) or ''),
            'source_file': tender.get('source_file', tender.get('Source File', '')),
            'ocr_confidence': float(tender.get('ocr_confidence', tender.get('OCR Confidence', 0)) or 0),
            'has_arabic': bool(tender.get('has_arabic', tender.get('Has Arabic', False))),
            'ai_enhanced': bool(tender.get('ai_enhanced', False)),
            'ai_provider': tender.get('ai_provider', ''),
            'extraction_date': str(tender.get('extraction_date', tender.get('Extraction Date', '')) or ''),
            'items': [
                {
                    'item_number': str(i.get('item_number', i.get('Item #', '')) or ''),
                    'description': str(i.get('description', i.get('Item Description', '')) or ''),
                    'quantity': str(i.get('quantity', i.get('Quantity', '')) or ''),
                    'unit': str(i.get('unit', i.get('Unit', 'PCS')) or 'PCS'),
                    'ai_cleaned': bool(i.get('ai_cleaned', False))
                }
                for i in tender.get('items', [])
                if i.get('description', i.get('Item Description', ''))
            ]
        }

    def _save_cache(self):
        """Save tenders to cache file"""
        try:
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'tenders': self.tenders,
                    'updated_at': datetime.now().isoformat(),
                    'count': len(self.tenders)
                }, f, indent=2, ensure_ascii=False, default=str)
            logger.info(f"Saved {len(self.tenders)} tenders to cache")
        except Exception as e:
            logger.error(f"Failed to save cache: {e}")

    # ========================================================================
    # PUBLIC API
    # ========================================================================

    def get_all_tenders(self, page: int = 1, limit: int = 50, search: str = None, department: str = None) -> Dict:
        """Get paginated list of tenders with filtering"""
        filtered = list(self.tenders.values())

        # Apply filters
        if search:
            search_lower = search.lower()
            filtered = [t for t in filtered if
                       search_lower in t.get('reference', '').lower() or
                       search_lower in t.get('title', '').lower() or
                       any(search_lower in i.get('description', '').lower() for i in t.get('items', []))]

        if department:
            filtered = [t for t in filtered if department.lower() in t.get('department', '').lower()]

        # Sort by reference
        filtered.sort(key=lambda x: x.get('reference', ''), reverse=True)

        # Paginate
        total = len(filtered)
        start = (page - 1) * limit
        end = start + limit

        return {
            'tenders': filtered[start:end],
            'total': total,
            'page': page,
            'limit': limit,
            'pages': (total + limit - 1) // limit
        }

    def get_tender(self, reference: str) -> Optional[Dict]:
        """Get a single tender by reference"""
        return self.tenders.get(reference)

    def get_dashboard_stats(self) -> Dict:
        """Get dashboard statistics"""
        tenders = list(self.tenders.values())

        if not tenders:
            return {
                'total_tenders': 0,
                'total_items': 0,
                'unique_departments': 0,
                'ai_enhanced_count': 0,
                'avg_confidence': 0,
                'items_per_tender': 0,
                'department_breakdown': {},
                'recent_tenders': [],
                'top_items': [],
                'ai_stats': self.ai_engine.get_stats()
            }

        total_items = sum(len(t.get('items', [])) for t in tenders)
        departments = {}
        for t in tenders:
            dept = t.get('department', 'Unknown') or 'Unknown'
            departments[dept] = departments.get(dept, 0) + 1

        confidences = [t.get('ocr_confidence', 0) for t in tenders if t.get('ocr_confidence')]
        avg_conf = sum(confidences) / len(confidences) if confidences else 0

        ai_enhanced = sum(1 for t in tenders if t.get('ai_enhanced'))

        # Recent tenders (last 10)
        sorted_tenders = sorted(tenders, key=lambda x: x.get('reference', ''), reverse=True)
        recent = [{'reference': t['reference'], 'title': t.get('title', ''), 'items': len(t.get('items', []))} for t in sorted_tenders[:10]]

        # Top items by frequency
        item_counts = {}
        for t in tenders:
            for item in t.get('items', []):
                desc = item.get('description', '')[:50]
                if desc and len(desc) > 5:
                    item_counts[desc] = item_counts.get(desc, 0) + 1

        top_items = sorted(item_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        top_items = [{'description': k, 'count': v} for k, v in top_items]

        return {
            'total_tenders': len(tenders),
            'total_items': total_items,
            'unique_departments': len(departments),
            'ai_enhanced_count': ai_enhanced,
            'avg_confidence': round(avg_conf, 1),
            'items_per_tender': round(total_items / len(tenders), 1) if tenders else 0,
            'department_breakdown': departments,
            'recent_tenders': recent,
            'top_items': top_items,
            'ai_stats': self.ai_engine.get_stats()
        }

    def enhance_tender(self, reference: str) -> Dict:
        """Apply AI enhancement to a single tender"""
        tender = self.tenders.get(reference)
        if not tender:
            return {'error': 'Tender not found'}

        enhanced = tender.copy()
        items_cleaned = 0

        # Clean each item description
        new_items = []
        for item in tender.get('items', []):
            desc = item.get('description', '')
            if not desc or len(desc) < 5:
                continue

            # Check if looks like noise
            is_noise = (
                desc.isdigit() or
                len(desc.split()) < 2 or
                re.match(r'^[\d\s.,-]+$', desc)
            )
            if is_noise:
                continue

            cleaned_desc, was_modified = self.ai_engine.clean_description(desc)
            if cleaned_desc == "INVALID":
                continue

            new_item = item.copy()
            if was_modified:
                new_item['description'] = cleaned_desc
                new_item['original_description'] = desc
                new_item['ai_cleaned'] = True
                items_cleaned += 1

            new_items.append(new_item)

        enhanced['items'] = new_items
        enhanced['ai_enhanced'] = True
        enhanced['ai_provider'] = 'gemini+groq'
        enhanced['items_cleaned'] = items_cleaned

        # Update cache
        self.tenders[reference] = enhanced
        self._save_cache()

        return enhanced

    def enhance_batch(self, references: List[str] = None, limit: int = 10) -> Dict:
        """Enhance multiple tenders with AI"""
        if references:
            to_enhance = [self.tenders[ref] for ref in references if ref in self.tenders]
        else:
            # Get tenders that haven't been enhanced yet
            to_enhance = [t for t in self.tenders.values() if not t.get('ai_enhanced')][:limit]

        results = {
            'enhanced': 0,
            'items_cleaned': 0,
            'errors': 0
        }

        for tender in to_enhance:
            ref = tender.get('reference')
            try:
                result = self.enhance_tender(ref)
                if 'error' not in result:
                    results['enhanced'] += 1
                    results['items_cleaned'] += result.get('items_cleaned', 0)
            except Exception as e:
                results['errors'] += 1
                logger.error(f"Error enhancing {ref}: {e}")

        return results

    def search_items(self, query: str, limit: int = 50) -> List[Dict]:
        """Search across all tender items"""
        results = []
        query_lower = query.lower()

        for tender in self.tenders.values():
            for item in tender.get('items', []):
                desc = item.get('description', '').lower()
                if query_lower in desc:
                    results.append({
                        'tender_reference': tender.get('reference'),
                        'department': tender.get('department'),
                        'item_number': item.get('item_number'),
                        'description': item.get('description'),
                        'quantity': item.get('quantity'),
                        'unit': item.get('unit')
                    })
                    if len(results) >= limit:
                        return results

        return results

    def get_departments(self) -> List[str]:
        """Get list of unique departments"""
        departments = set()
        for tender in self.tenders.values():
            dept = tender.get('department', '')
            if dept:
                departments.add(dept)
        return sorted(departments)

    def export_to_excel(self, references: List[str] = None) -> str:
        """Export tenders to Excel file"""
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl not available")

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Tenders"

        # Headers
        headers = ['Reference', 'Department', 'Title', 'Closing Date', 'Item #',
                   'Description', 'Quantity', 'Unit', 'AI Cleaned', 'OCR Confidence']
        ws.append(headers)

        # Data
        tenders_to_export = [self.tenders[ref] for ref in references] if references else self.tenders.values()

        for tender in tenders_to_export:
            for item in tender.get('items', []):
                ws.append([
                    tender.get('reference', ''),
                    tender.get('department', ''),
                    tender.get('title', ''),
                    tender.get('closing_date', ''),
                    item.get('item_number', ''),
                    item.get('description', ''),
                    item.get('quantity', ''),
                    item.get('unit', ''),
                    'Yes' if item.get('ai_cleaned') else 'No',
                    tender.get('ocr_confidence', 0)
                ])

        # Save
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_dir = Path(CONFIG["ai_enhanced_dir"])
        export_dir.mkdir(parents=True, exist_ok=True)
        filepath = export_dir / f"export_{timestamp}.xlsx"
        wb.save(filepath)

        return str(filepath)


# ============================================================================
# SINGLETON INSTANCE
# ============================================================================

def get_tender_manager() -> TenderManager:
    """Get the singleton TenderManager instance"""
    return TenderManager()


def get_ai_engine() -> AIEngine:
    """Get the singleton AIEngine instance"""
    return AIEngine()
