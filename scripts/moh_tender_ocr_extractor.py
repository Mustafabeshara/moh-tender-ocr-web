# -*- coding: utf-8 -*-
"""
OCR-Enhanced MOH Tender Extractor

This module provides OCR capabilities for extracting structured data from
MOH Kuwait biomedical engineering tender PDFs, including:
- Reference Number
- Closing Date
- Items (line by line)
- Description
- Quantity
- Specifications (from V2 files)

Author: MOH Tender Automation
Version: 1.0.0
"""

import os
import sys
import re
import json
import logging
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, field, asdict

# ============================================================================
# CONFIGURATION
# ============================================================================

CONFIG = {
    "base_folder": os.path.expanduser("~/Desktop/MOH Tenders"),
    "biomedical_folder": "Biomedical_Engineering",
    "medical_store_folder": "Medical_Store",
    "output_folder": "OCR_Extracted_Data",
    "log_folder": "logs",
    "tesseract_path": "/opt/homebrew/bin/tesseract",
    "languages": ["eng", "ara"],
    "dpi": 300,
    "max_pages": 10,
    "debug_mode": False,
    "save_intermediate_images": False,
}

# ============================================================================
# DEPENDENCY INSTALLATION
# ============================================================================

def install_dependencies():
    """Install required Python packages."""
    packages = ["pytesseract", "pdf2image", "Pillow", "tqdm", "openpyxl"]

    for package in packages:
        try:
            __import__(package.replace("-", "_").lower())
            print(f"✅ {package} already installed")
        except ImportError:
            print(f"📦 Installing {package}...")
            try:
                subprocess.check_call([
                    sys.executable, "-m", "pip", "install",
                    package, "--break-system-packages", "-q"
                ])
                print(f"✅ {package} installed successfully")
            except subprocess.CalledProcessError as e:
                print(f"❌ Failed to install {package}: {e}")
                return False
    return True

def check_tesseract():
    """Check if Tesseract OCR is installed."""
    tesseract_paths = [
        "/opt/homebrew/bin/tesseract",
        "/usr/local/bin/tesseract",
        "/usr/bin/tesseract",
        "tesseract",
    ]

    for path in tesseract_paths:
        try:
            result = subprocess.run([path, "--version"], capture_output=True, text=True)
            if result.returncode == 0:
                version = result.stdout.split('\n')[0]
                print(f"✅ Tesseract found: {path}")
                print(f"   Version: {version}")
                CONFIG["tesseract_path"] = path
                return True
        except (FileNotFoundError, subprocess.SubprocessError):
            continue

    print("❌ Tesseract OCR not found!")
    print("   Install with: brew install tesseract tesseract-lang")
    return False

def check_poppler():
    """Check if Poppler (for pdf2image) is installed."""
    try:
        result = subprocess.run(["pdftoppm", "-v"], capture_output=True, text=True)
        print("✅ Poppler found (required for pdf2image)")
        return True
    except FileNotFoundError:
        print("❌ Poppler not found!")
        print("   Install with: brew install poppler")
        return False

# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class TenderItem:
    """Represents a single item in a tender."""
    item_number: str = ""
    description: str = ""
    quantity: str = ""
    unit: str = ""
    specifications: str = ""
    language: str = "unknown"
    has_arabic: bool = False

    def to_dict(self) -> Dict:
        return asdict(self)

@dataclass
class TenderData:
    """Represents extracted data from a tender."""
    reference_number: str = ""
    title: str = ""
    closing_date: str = ""
    posting_date: str = ""
    department: str = ""
    items: List[TenderItem] = field(default_factory=list)
    specifications_text: str = ""
    raw_text: str = ""
    ocr_confidence: float = 0.0
    source_files: List[str] = field(default_factory=list)
    extraction_timestamp: str = ""
    extraction_method: str = "ocr"
    language: str = "unknown"
    has_arabic_content: bool = False
    errors: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict:
        return {
            "reference_number": self.reference_number,
            "title": self.title,
            "closing_date": self.closing_date,
            "posting_date": self.posting_date,
            "department": self.department,
            "items": [item.to_dict() for item in self.items],
            "specifications_text": self.specifications_text,
            "items_count": len(self.items),
            "ocr_confidence": self.ocr_confidence,
            "source_files": self.source_files,
            "extraction_timestamp": self.extraction_timestamp,
            "extraction_method": self.extraction_method,
            "language": self.language,
            "has_arabic_content": self.has_arabic_content,
            "errors": self.errors,
        }

# ============================================================================
# LOGGING SETUP
# ============================================================================

def setup_logging():
    """Configure logging for the OCR extractor."""
    log_folder = os.path.join(CONFIG["base_folder"], CONFIG["log_folder"])
    os.makedirs(log_folder, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_folder, f"ocr_extraction_{timestamp}.log")

    logging.basicConfig(
        level=logging.DEBUG if CONFIG["debug_mode"] else logging.INFO,
        format='%(asctime)s | %(levelname)8s | %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    return logging.getLogger(__name__)

# ============================================================================
# OCR ENGINE
# ============================================================================

class OCREngine:
    """OCR engine using Tesseract for PDF text extraction."""

    def __init__(self, logger=None):
        self.logger = logger or logging.getLogger(__name__)
        self.tesseract_path = CONFIG["tesseract_path"]

        import pytesseract
        from pdf2image import convert_from_path
        from PIL import Image

        self.pytesseract = pytesseract
        self.convert_from_path = convert_from_path
        self.Image = Image

        pytesseract.pytesseract.tesseract_cmd = self.tesseract_path

        self.logger.info(f"OCR Engine initialized with Tesseract at {self.tesseract_path}")

    def extract_text_from_pdf(self, pdf_path: str, languages: List[str] = None) -> Tuple[str, float]:
        """Extract text from PDF using OCR."""
        if languages is None:
            languages = CONFIG["languages"]

        lang_str = "+".join(languages)

        self.logger.info(f"📄 Processing PDF: {os.path.basename(pdf_path)}")

        if not os.path.exists(pdf_path):
            self.logger.error(f"PDF not found: {pdf_path}")
            return "", 0.0

        try:
            self.logger.info(f"   Converting to images (DPI: {CONFIG['dpi']})...")
            images = self.convert_from_path(
                pdf_path,
                dpi=CONFIG["dpi"],
                first_page=1,
                last_page=CONFIG["max_pages"]
            )

            self.logger.info(f"   Converted {len(images)} pages")

            all_text = []
            confidences = []

            for i, image in enumerate(images, 1):
                self.logger.info(f"   OCR processing page {i}/{len(images)}...")

                processed_image = self._preprocess_image(image)

                ocr_data = self.pytesseract.image_to_data(
                    processed_image,
                    lang=lang_str,
                    output_type=self.pytesseract.Output.DICT,
                    config='--psm 6 --oem 3'
                )

                page_text = []
                page_confidences = []

                for j, word in enumerate(ocr_data['text']):
                    conf = int(ocr_data['conf'][j])
                    if conf > 0 and word.strip():
                        page_text.append(word)
                        page_confidences.append(conf)

                if page_confidences:
                    avg_conf = sum(page_confidences) / len(page_confidences)
                    confidences.append(avg_conf)

                all_text.append(" ".join(page_text))

            full_text = "\n\n--- PAGE BREAK ---\n\n".join(all_text)
            avg_confidence = sum(confidences) / len(confidences) if confidences else 0.0

            self.logger.info(f"   ✅ Extracted {len(full_text)} characters, confidence: {avg_confidence:.1f}%")

            return full_text, avg_confidence

        except Exception as e:
            self.logger.error(f"   ❌ OCR failed: {e}")
            return "", 0.0

    def _preprocess_image(self, image):
        """Preprocess image for better OCR accuracy."""
        from PIL import ImageEnhance, ImageFilter

        if image.mode != 'L':
            image = image.convert('L')

        enhancer = ImageEnhance.Contrast(image)
        image = enhancer.enhance(1.5)

        image = image.filter(ImageFilter.SHARPEN)

        threshold = 128
        image = image.point(lambda p: 255 if p > threshold else 0)

        return image

# ============================================================================
# TEXT PARSERS
# ============================================================================

class TenderParser:
    """Parser for extracting structured data from OCR text."""

    def __init__(self, logger=None):
        self.logger = logger or logging.getLogger(__name__)

    def detect_language(self, text: str) -> Dict:
        """Detect if text contains Arabic content."""
        arabic_pattern = r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\uFB50-\uFDFF\uFE70-\uFEFF]'
        arabic_chars = len(re.findall(arabic_pattern, text))
        total_chars = len(re.sub(r'\s+', '', text))

        if total_chars == 0:
            return {'language': 'unknown', 'arabic_ratio': 0, 'has_arabic': False}

        arabic_ratio = arabic_chars / total_chars

        if arabic_ratio > 0.3:
            primary_language = 'arabic'
        elif arabic_ratio > 0.05:
            primary_language = 'mixed'
        else:
            primary_language = 'english'

        return {
            'language': primary_language,
            'arabic_ratio': arabic_ratio,
            'has_arabic': arabic_chars > 0
        }

    def extract_reference_number(self, text: str, filename: str = "") -> str:
        """Extract tender reference number from text or filename."""
        if filename:
            filename_match = re.search(r'(\d{1,2}[A-Z]{2,3}\d{2,4})', filename)
            if filename_match:
                return filename_match.group(1)

        patterns = [
            r'(?:Tender\s*(?:No\.?|Number|Ref\.?)?[:\s]*)?(\d{1,2}[A-Z]{2,3}\d{2,4})',
            r'(?:Reference[:\s]*)?(\d{1,2}[A-Z]{2,3}\d{2,4})',
            r'(?:رقم\s*المناقصة[:\s]*)?(\d{1,2}[A-Z]{2,3}\d{2,4})',
            r'\b(\d{1,2}(?:TN|LB|AL|EQ|LS|MA|PS|PT|TE|TS|IC|RC|BM)\d{2,4})\b',
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).upper()

        return ""

    def extract_closing_date(self, text: str) -> str:
        """Extract closing date from tender text."""
        patterns = [
            r'(?:Closing\s*Date|Close\s*Date|Last\s*Date|Deadline)[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
            r'(?:تاريخ\s*الاغلاق|تاريخ\s*الانتهاء)[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
            r'(?:close[sd]?\s*(?:on|by)?|deadline)[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
            r'(\d{2}/\d{2}/\d{4})',
        ]

        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                date_str = matches[-1] if len(matches) > 1 else matches[0]
                date_str = date_str.replace('-', '/').replace('.', '/')
                return date_str

        return ""

    def extract_posting_date(self, text: str) -> str:
        """Extract posting/publication date from tender text."""
        patterns = [
            r'(?:Posted|Published|Posted\s*Date|Publication\s*Date)[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
            r'(?:تاريخ\s*النشر|تاريخ\s*الاعلان)[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                date_str = match.group(1)
                date_str = date_str.replace('-', '/').replace('.', '/')
                return date_str

        all_dates = re.findall(r'(\d{2}/\d{2}/\d{4})', text)
        if all_dates:
            return all_dates[0]

        return ""

    def extract_items(self, text: str) -> List[TenderItem]:
        """Extract items with descriptions and quantities from tender text."""
        items = []
        lines = text.split('\n')

        item_patterns = [
            r'^[\s]*(\d+)[.):\-\s]+([^0-9]+?)\s*[-–:]\s*(\d+(?:[.,]\d+)?)\s*(pieces?|pcs?|units?|each|nos?|sets?|qty|قطعة|وحدة|عدد)?',
            r'^[\s]*(\d+)\s+([^\t]+?)\t+\s*(\d+(?:[.,]\d+)?)\s*(pieces?|pcs?|units?|each|nos?|sets?)?',
            r'^[\s]*(\d+)\s*\|\s*([^|]+?)\s*\|\s*(\d+(?:[.,]\d+)?)',
            r'^[\s]*(\d+)[.)\s]+(.+?)\s+(\d+)\s*(pieces?|pcs?|units?|each|nos?|sets?|qty)?[\s]*$',
        ]

        current_item_number = 0

        for line in lines:
            line = line.strip()
            if len(line) < 5:
                continue

            for pattern in item_patterns:
                match = re.match(pattern, line, re.IGNORECASE | re.UNICODE)
                if match:
                    groups = match.groups()

                    item_no = groups[0]
                    description = groups[1].strip() if len(groups) > 1 else ""
                    quantity = groups[2] if len(groups) > 2 else ""
                    unit = groups[3] if len(groups) > 3 and groups[3] else "units"

                    description = re.sub(r'\s+', ' ', description).strip()
                    description = re.sub(r'^[\d\s\-\.]+', '', description)

                    if quantity:
                        quantity = re.sub(r'[^\d.,]', '', quantity)

                    if description and len(description) > 3:
                        lang_info = self.detect_language(description)

                        items.append(TenderItem(
                            item_number=str(item_no),
                            description=description[:500],
                            quantity=quantity,
                            unit=unit.lower() if unit else "units",
                            language=lang_info['language'],
                            has_arabic=lang_info['has_arabic']
                        ))
                        current_item_number = int(item_no)
                    break

            if not items or current_item_number == 0:
                qty_match = re.search(
                    r'(.{10,}?)\s+(\d+)\s*(pieces?|pcs?|units?|each|nos?|sets?|qty|قطعة|وحدة)\s*$',
                    line,
                    re.IGNORECASE
                )
                if qty_match:
                    description = qty_match.group(1).strip()
                    quantity = qty_match.group(2)
                    unit = qty_match.group(3)

                    if len(description) > 5:
                        current_item_number += 1
                        lang_info = self.detect_language(description)

                        items.append(TenderItem(
                            item_number=str(current_item_number),
                            description=description[:500],
                            quantity=quantity,
                            unit=unit.lower(),
                            language=lang_info['language'],
                            has_arabic=lang_info['has_arabic']
                        ))

        unique_items = []
        seen_descriptions = set()

        for item in items:
            desc_key = item.description.lower()[:100]
            if desc_key not in seen_descriptions:
                seen_descriptions.add(desc_key)
                unique_items.append(item)

        self.logger.info(f"   Found {len(unique_items)} unique items")
        return unique_items

    def extract_specifications(self, text: str) -> str:
        """Extract specifications section from V2 documents."""
        spec_markers = [
            r'(?:Technical\s*)?Specifications?[:\s]*(.+?)(?:Terms|Conditions|Notes|\Z)',
            r'(?:المواصفات\s*الفنية|المواصفات)[:\s]*(.+?)(?:الشروط|ملاحظات|\Z)',
            r'Requirements?[:\s]*(.+?)(?:Terms|Notes|\Z)',
            r'Description[:\s]*(.+?)(?:Quantity|Terms|\Z)',
        ]

        for pattern in spec_markers:
            match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
            if match:
                specs = match.group(1).strip()
                specs = re.sub(r'\s+', ' ', specs)
                if len(specs) > 50:
                    return specs[:2000]

        paragraphs = re.split(r'\n\s*\n', text)
        for para in paragraphs:
            para = para.strip()
            if len(para) > 100:
                return para[:2000]

        return ""

# ============================================================================
# MAIN EXTRACTOR
# ============================================================================

class TenderExtractor:
    """Main class for extracting data from MOH tender PDFs."""

    def __init__(self, logger=None):
        self.logger = logger or setup_logging()
        self.ocr_engine = OCREngine(self.logger)
        self.parser = TenderParser(self.logger)
        self.results = []

        self.output_folder = os.path.join(CONFIG["base_folder"], CONFIG["output_folder"])
        os.makedirs(self.output_folder, exist_ok=True)

    def process_tender_set(self, reference: str, pdf_files: Dict[str, str]) -> TenderData:
        """Process a set of PDFs for a single tender (V1, V2, V3)."""
        self.logger.info(f"\n{'='*60}")
        self.logger.info(f"Processing Tender: {reference}")
        self.logger.info(f"{'='*60}")

        tender_data = TenderData(
            reference_number=reference,
            extraction_timestamp=datetime.now().isoformat(),
            source_files=list(pdf_files.values())
        )

        if 'v1' in pdf_files and pdf_files['v1']:
            self.logger.info(f"\n📄 Processing V1 (Main Tender)...")
            v1_text, v1_conf = self.ocr_engine.extract_text_from_pdf(pdf_files['v1'])

            if v1_text:
                tender_data.closing_date = self.parser.extract_closing_date(v1_text)
                tender_data.posting_date = self.parser.extract_posting_date(v1_text)
                tender_data.items = self.parser.extract_items(v1_text)
                tender_data.ocr_confidence = v1_conf

                lang_info = self.parser.detect_language(v1_text)
                tender_data.language = lang_info['language']
                tender_data.has_arabic_content = lang_info['has_arabic']

                tender_data.raw_text = v1_text[:5000]

                self.logger.info(f"   📅 Closing Date: {tender_data.closing_date}")
                self.logger.info(f"   📦 Items Found: {len(tender_data.items)}")

        if 'v2' in pdf_files and pdf_files['v2']:
            self.logger.info(f"\n📋 Processing V2 (Specifications)...")
            v2_text, v2_conf = self.ocr_engine.extract_text_from_pdf(pdf_files['v2'])

            if v2_text:
                tender_data.specifications_text = self.parser.extract_specifications(v2_text)

                v2_items = self.parser.extract_items(v2_text)

                for i, item in enumerate(tender_data.items):
                    if i < len(v2_items):
                        item.specifications = v2_items[i].description

                self.logger.info(f"   📝 Specifications extracted: {len(tender_data.specifications_text)} chars")

        if 'v3' in pdf_files and pdf_files['v3']:
            self.logger.info(f"\n📝 Processing V3 (Extra Notes)...")
            v3_text, _ = self.ocr_engine.extract_text_from_pdf(pdf_files['v3'])

            if v3_text:
                additional_specs = self.parser.extract_specifications(v3_text)
                if additional_specs:
                    tender_data.specifications_text += f"\n\n[Additional Notes]\n{additional_specs}"

        return tender_data

    def process_folder(self, folder_path: str, department: str = "Biomedical Engineering") -> List[TenderData]:
        """Process all tender PDFs in a folder."""
        self.logger.info(f"\n{'#'*60}")
        self.logger.info(f"# Processing Folder: {folder_path}")
        self.logger.info(f"# Department: {department}")
        self.logger.info(f"{'#'*60}")

        if not os.path.exists(folder_path):
            self.logger.error(f"Folder not found: {folder_path}")
            return []

        tender_groups = {}

        for filename in os.listdir(folder_path):
            if not filename.endswith('.pdf'):
                continue

            ref_match = re.search(r'^(\d{1,2}[A-Z]{2,3}\d{2,4})', filename)
            if not ref_match:
                continue

            reference = ref_match.group(1)

            if reference not in tender_groups:
                tender_groups[reference] = {'v1': None, 'v2': None, 'v3': None}

            file_path = os.path.join(folder_path, filename)
            if '_V1' in filename:
                tender_groups[reference]['v1'] = file_path
            elif '_V2' in filename:
                tender_groups[reference]['v2'] = file_path
            elif '_V3' in filename:
                tender_groups[reference]['v3'] = file_path

        self.logger.info(f"Found {len(tender_groups)} tender references to process")

        results = []

        try:
            from tqdm import tqdm
            iterator = tqdm(tender_groups.items(), desc="Processing tenders")
        except ImportError:
            iterator = tender_groups.items()
            print(f"Processing {len(tender_groups)} tenders...")

        for reference, files in iterator:
            try:
                tender_data = self.process_tender_set(reference, files)
                tender_data.department = department
                results.append(tender_data)
            except Exception as e:
                self.logger.error(f"Error processing {reference}: {e}")
                tender_data = TenderData(
                    reference_number=reference,
                    department=department,
                    errors=[str(e)],
                    extraction_timestamp=datetime.now().isoformat()
                )
                results.append(tender_data)

        self.results = results
        return results

    def save_results(self, output_format: str = "all") -> Dict[str, str]:
        """Save extraction results to files."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_files = {}

        if output_format in ['json', 'all']:
            json_path = os.path.join(self.output_folder, f"ocr_extracted_{timestamp}.json")

            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(
                    {
                        "extraction_date": datetime.now().isoformat(),
                        "total_tenders": len(self.results),
                        "total_items": sum(len(t.items) for t in self.results),
                        "tenders": [t.to_dict() for t in self.results]
                    },
                    f,
                    ensure_ascii=False,
                    indent=2
                )

            output_files['json'] = json_path
            self.logger.info(f"📄 JSON saved: {json_path}")

        if output_format in ['excel', 'all']:
            excel_path = os.path.join(self.output_folder, f"ocr_extracted_{timestamp}.xlsx")
            self._save_excel(excel_path)
            output_files['excel'] = excel_path
            self.logger.info(f"📊 Excel saved: {excel_path}")

        return output_files

    def _save_excel(self, excel_path: str):
        """Save results to Excel with multiple sheets."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        summary_sheet = wb.create_sheet("Summary")
        self._create_summary_sheet(summary_sheet)

        tenders_sheet = wb.create_sheet("Tenders")
        self._create_tenders_sheet(tenders_sheet)

        items_sheet = wb.create_sheet("Items")
        self._create_items_sheet(items_sheet)

        specs_sheet = wb.create_sheet("Specifications")
        self._create_specs_sheet(specs_sheet)

        wb.save(excel_path)

    def _create_summary_sheet(self, sheet):
        """Create summary sheet."""
        from openpyxl.styles import Font, PatternFill

        sheet['A1'] = 'MOH Tender OCR Extraction Report'
        sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color='2E8B57', fill_type='solid')
        sheet.merge_cells('A1:D1')

        stats = [
            ('Extraction Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Total Tenders Processed:', len(self.results)),
            ('Total Items Extracted:', sum(len(t.items) for t in self.results)),
            ('Arabic Tenders:', len([t for t in self.results if t.has_arabic_content])),
            ('Tenders with Specifications:', len([t for t in self.results if t.specifications_text])),
        ]

        for row, (label, value) in enumerate(stats, 3):
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)

        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 40

    def _create_tenders_sheet(self, sheet):
        """Create tenders sheet."""
        from openpyxl.styles import Font, PatternFill
        import openpyxl

        headers = ['Reference', 'Closing Date', 'Posting Date', 'Items Count',
                   'Language', 'OCR Confidence', 'Has Specs', 'Department']

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')

        for row, tender in enumerate(self.results, 2):
            sheet.cell(row=row, column=1, value=tender.reference_number)
            sheet.cell(row=row, column=2, value=tender.closing_date)
            sheet.cell(row=row, column=3, value=tender.posting_date)
            sheet.cell(row=row, column=4, value=len(tender.items))
            sheet.cell(row=row, column=5, value=tender.language)
            sheet.cell(row=row, column=6, value=f"{tender.ocr_confidence:.1f}%")
            sheet.cell(row=row, column=7, value="Yes" if tender.specifications_text else "No")
            sheet.cell(row=row, column=8, value=tender.department)

        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    def _create_items_sheet(self, sheet):
        """Create items sheet with all extracted items."""
        from openpyxl.styles import Font, PatternFill
        import openpyxl

        headers = ['Tender Ref', 'Item #', 'Description', 'Quantity', 'Unit',
                   'Specifications', 'Language', 'Has Arabic']

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='E74C3C', fill_type='solid')

        row = 2
        for tender in self.results:
            for item in tender.items:
                sheet.cell(row=row, column=1, value=tender.reference_number)
                sheet.cell(row=row, column=2, value=item.item_number)
                sheet.cell(row=row, column=3, value=item.description)
                sheet.cell(row=row, column=4, value=item.quantity)
                sheet.cell(row=row, column=5, value=item.unit)
                sheet.cell(row=row, column=6, value=item.specifications)
                sheet.cell(row=row, column=7, value=item.language)
                sheet.cell(row=row, column=8, value="Yes" if item.has_arabic else "No")
                row += 1

        widths = [12, 8, 60, 10, 10, 50, 10, 10]
        for col, width in enumerate(widths, 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    def _create_specs_sheet(self, sheet):
        """Create specifications sheet."""
        from openpyxl.styles import Font, PatternFill
        import openpyxl

        headers = ['Tender Ref', 'Closing Date', 'Specifications']

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='8E44AD', fill_type='solid')

        for row, tender in enumerate(self.results, 2):
            sheet.cell(row=row, column=1, value=tender.reference_number)
            sheet.cell(row=row, column=2, value=tender.closing_date)
            sheet.cell(row=row, column=3, value=tender.specifications_text[:1000] if tender.specifications_text else "N/A")

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 100

# ============================================================================
# TEST FUNCTIONS
# ============================================================================

def test_single_pdf(pdf_path: str):
    """Test OCR extraction on a single PDF."""
    print(f"\n{'='*60}")
    print(f"Testing OCR on: {os.path.basename(pdf_path)}")
    print(f"{'='*60}")

    logger = setup_logging()

    ocr_engine = OCREngine(logger)
    parser = TenderParser(logger)

    text, confidence = ocr_engine.extract_text_from_pdf(pdf_path)

    if not text:
        print("No text extracted from PDF")
        return None

    print(f"\nExtracted Text Preview (first 1000 chars):")
    print("-" * 40)
    print(text[:1000])
    print("-" * 40)

    print("\nParsed Data:")

    reference = parser.extract_reference_number(text, os.path.basename(pdf_path))
    print(f"   Reference Number: {reference}")

    closing_date = parser.extract_closing_date(text)
    print(f"   Closing Date: {closing_date}")

    posting_date = parser.extract_posting_date(text)
    print(f"   Posting Date: {posting_date}")

    items = parser.extract_items(text)
    print(f"   Items Found: {len(items)}")

    for i, item in enumerate(items[:5], 1):
        print(f"      {i}. {item.description[:50]}... | Qty: {item.quantity} {item.unit}")

    if len(items) > 5:
        print(f"      ... and {len(items) - 5} more items")

    lang_info = parser.detect_language(text)
    print(f"   Language: {lang_info['language']} (Arabic ratio: {lang_info['arabic_ratio']:.1%})")
    print(f"   OCR Confidence: {confidence:.1f}%")

    return {
        'reference': reference,
        'closing_date': closing_date,
        'items': items,
        'confidence': confidence,
        'text_preview': text[:2000]
    }

def test_tender_set(reference: str):
    """Test extraction on a full tender set (V1, V2, V3)."""
    base_folder = os.path.join(CONFIG["base_folder"], CONFIG["biomedical_folder"])

    pdf_files = {'v1': None, 'v2': None, 'v3': None}

    for filename in os.listdir(base_folder):
        if filename.startswith(reference) and filename.endswith('.pdf'):
            file_path = os.path.join(base_folder, filename)
            if '_V1' in filename:
                pdf_files['v1'] = file_path
            elif '_V2' in filename:
                pdf_files['v2'] = file_path
            elif '_V3' in filename:
                pdf_files['v3'] = file_path

    print(f"\n{'='*60}")
    print(f"Testing Full Tender Set: {reference}")
    print(f"{'='*60}")
    print(f"   V1: {os.path.basename(pdf_files['v1']) if pdf_files['v1'] else 'Not found'}")
    print(f"   V2: {os.path.basename(pdf_files['v2']) if pdf_files['v2'] else 'Not found'}")
    print(f"   V3: {os.path.basename(pdf_files['v3']) if pdf_files['v3'] else 'Not found'}")

    logger = setup_logging()
    extractor = TenderExtractor(logger)

    tender_data = extractor.process_tender_set(reference, pdf_files)

    print(f"\nExtraction Results:")
    print(f"   Reference: {tender_data.reference_number}")
    print(f"   Closing Date: {tender_data.closing_date}")
    print(f"   Posting Date: {tender_data.posting_date}")
    print(f"   Items: {len(tender_data.items)}")
    print(f"   Specifications: {len(tender_data.specifications_text)} chars")
    print(f"   Language: {tender_data.language}")
    print(f"   Confidence: {tender_data.ocr_confidence:.1f}%")

    return tender_data

def run_full_extraction(department: str = "Biomedical Engineering", limit: int = None):
    """Run full extraction on all PDFs in a department folder."""
    print(f"\n{'#'*60}")
    print(f"# MOH Tender OCR Extraction - Full Run")
    print(f"# Department: {department}")
    print(f"# Limit: {limit if limit else 'All'}")
    print(f"{'#'*60}")

    logger = setup_logging()
    extractor = TenderExtractor(logger)

    if department == "Biomedical Engineering":
        folder_path = os.path.join(CONFIG["base_folder"], CONFIG["biomedical_folder"])
    else:
        folder_path = os.path.join(CONFIG["base_folder"], CONFIG["medical_store_folder"])

    results = extractor.process_folder(folder_path, department)

    if limit:
        results = results[:limit]
        extractor.results = results

    output_files = extractor.save_results("all")

    print(f"\n{'='*60}")
    print(f"Extraction Complete!")
    print(f"   Tenders Processed: {len(results)}")
    print(f"   Total Items: {sum(len(t.items) for t in results)}")
    print(f"   JSON Output: {output_files.get('json', 'N/A')}")
    print(f"   Excel Output: {output_files.get('excel', 'N/A')}")
    print(f"{'='*60}")

    return results

# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    """Main entry point."""
    print("\n" + "="*60)
    print("MOH Tender OCR Extractor v1.0.0")
    print("="*60)

    print("\nChecking dependencies...")

    if not install_dependencies():
        print("Failed to install dependencies")
        return False

    if not check_tesseract():
        print("Tesseract OCR not found")
        print("   Install with: brew install tesseract tesseract-lang")
        return False

    if not check_poppler():
        print("Poppler not found")
        print("   Install with: brew install poppler")
        return False

    print("\nAll dependencies ready!")
    print("\nAvailable functions:")
    print("   - test_single_pdf(pdf_path)     : Test OCR on single PDF")
    print("   - test_tender_set('5TN123')     : Test full tender set")
    print("   - run_full_extraction()         : Process all tenders")
    print("   - run_full_extraction(limit=10) : Process first 10 tenders")

    return True

if __name__ == "__main__":
    if main():
        print("\nReady! Run test_single_pdf() or run_full_extraction() to begin.")
