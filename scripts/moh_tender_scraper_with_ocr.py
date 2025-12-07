# -*- coding: utf-8 -*-
"""
MOH Tender Scraper with Integrated OCR
=======================================
Complete solution that:
1. Scrapes tenders from MOH Kuwait website using Selenium
2. Downloads PDF files for each tender
3. Extracts text using Tesseract OCR (high accuracy for Arabic/English)
4. Parses items, quantities, specifications
5. Exports to JSON and Excel

Requirements:
    pip install selenium beautifulsoup4 chromedriver-autoinstaller openpyxl pandas pytesseract pdf2image Pillow
    brew install tesseract tesseract-lang poppler (macOS)

Author: MOH Tender Analysis System
Date: December 2025
"""

import os
import sys
import subprocess
import time
import json
import re
import random
import requests
import threading
import hashlib
import sqlite3
import shutil
import logging
from datetime import datetime, timedelta
from urllib.parse import urljoin, urlparse
from collections import defaultdict
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, asdict
from pathlib import Path

# ============================================================================
# CONFIGURATION
# ============================================================================

CONFIG = {
    "base_folder": os.path.expanduser("~/Documents/MOH Tenders"),
    # Department-specific PDF folders (matching existing structure)
    "biomedical_folder": os.path.expanduser("~/Documents/MOH Tenders/Biomedical_Engineering"),
    "medical_store_folder": os.path.expanduser("~/Documents/MOH Tenders/Medical_Store"),
    "biomedical_shared_folder": os.path.expanduser("~/Documents/MOH Tenders/Biomedical_Engineering_Shared"),
    "medical_store_shared_folder": os.path.expanduser("~/Documents/MOH Tenders/Medical_Store_Shared"),
    # Output folders
    "output_folder": os.path.expanduser("~/Documents/MOH Tenders/scripts/scraping_results"),
    "ocr_output_folder": os.path.expanduser("~/Documents/MOH Tenders/ocr_results"),
    "log_file": os.path.expanduser("~/Documents/MOH Tenders/scraper_ocr.log"),
    # OCR settings
    "tesseract_lang": "eng+ara",
    "dpi": 300,
    "max_pages": 10,
    # Scraper settings
    "fast_mode": True,
    "headless_mode": True,
    "enable_ocr": True,  # Set to False to disable OCR
    "home_url": "https://www.moh.gov.kw/en/Pages/default.aspx",
}

# ============================================================================
# LOGGING SETUP
# ============================================================================

os.makedirs(os.path.dirname(CONFIG["log_file"]), exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    handlers=[
        logging.FileHandler(CONFIG["log_file"], encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================================
# AUTOMATED DEPENDENCY INSTALLATION
# ============================================================================

def install_and_import(package, import_name=None):
    """Check if a package is installed, install it if not, then import it."""
    import_name = import_name or package
    try:
        __import__(import_name)
    except ImportError:
        logger.info(f"'{package}' not found. Installing...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        except Exception as e:
            logger.error(f"Failed to install '{package}': {e}")
            sys.exit(1)
    return __import__(import_name)

# Install required packages
install_and_import('selenium')
install_and_import('bs4')
install_and_import('chromedriver_autoinstaller', 'chromedriver_autoinstaller')
install_and_import('openpyxl')
install_and_import('pandas')
install_and_import('PyPDF2')

# OCR packages (optional but recommended)
try:
    install_and_import('pytesseract')
    install_and_import('pdf2image')
    install_and_import('Pillow', 'PIL')
    OCR_AVAILABLE = True
except:
    OCR_AVAILABLE = False
    logger.warning("OCR packages not available. Install: pip install pytesseract pdf2image Pillow")
    logger.warning("Also install: brew install tesseract tesseract-lang poppler")

# Import libraries
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import chromedriver_autoinstaller
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import PyPDF2

if OCR_AVAILABLE:
    import pytesseract
    from pdf2image import convert_from_path
    from PIL import Image, ImageEnhance, ImageFilter

# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class TenderItem:
    """Represents a single item in a tender"""
    item_number: str
    description: str
    quantity: str
    unit: str = ""
    specifications: str = ""
    language: str = "unknown"
    has_arabic: bool = False

@dataclass
class TenderData:
    """Represents extracted tender data"""
    reference: str
    closing_date: str = ""
    posting_date: str = ""
    department: str = ""
    title: str = ""
    items: List[TenderItem] = None
    specifications: str = ""
    raw_text: str = ""
    ocr_confidence: float = 0.0
    source_files: List[str] = None
    extraction_date: str = ""
    has_arabic_content: bool = False
    amount: str = ""

    def __post_init__(self):
        if self.items is None:
            self.items = []
        if self.source_files is None:
            self.source_files = []
        if not self.extraction_date:
            self.extraction_date = datetime.now().isoformat()

# ============================================================================
# DUPLICATE PREVENTION SYSTEM
# ============================================================================

class DuplicateMonitor:
    """Advanced duplicate detection and prevention system"""

    def __init__(self, base_folder):
        self.base_folder = base_folder
        self.db_path = os.path.join(base_folder, 'duplicate_tracker.db')
        self.init_database()
        self.file_hashes = {}
        self.tender_refs = set()

    def init_database(self):
        """Initialize SQLite database for tracking downloads"""
        try:
            os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS downloaded_files (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tender_ref TEXT NOT NULL,
                    file_path TEXT NOT NULL,
                    file_hash TEXT NOT NULL,
                    file_size INTEGER,
                    download_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    department TEXT,
                    version TEXT,
                    ocr_processed INTEGER DEFAULT 0,
                    UNIQUE(tender_ref, version)
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS tender_metadata (
                    tender_ref TEXT PRIMARY KEY,
                    title TEXT,
                    department TEXT,
                    posting_date TEXT,
                    closing_date TEXT,
                    amount TEXT,
                    ocr_confidence REAL,
                    items_count INTEGER,
                    last_checked TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            conn.commit()
            conn.close()
            logger.info(f"💾 Duplicate tracking database initialized: {self.db_path}")

        except Exception as e:
            logger.error(f"❌ Error initializing duplicate database: {e}")

    def calculate_file_hash(self, file_path):
        """Calculate SHA256 hash of a file"""
        try:
            hash_sha256 = hashlib.sha256()
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_sha256.update(chunk)
            return hash_sha256.hexdigest()
        except Exception as e:
            logger.warning(f"⚠️ Error calculating hash for {file_path}: {e}")
            return None

    def is_tender_duplicate(self, tender_ref, department):
        """Check if tender reference already exists"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT COUNT(*) FROM tender_metadata WHERE tender_ref = ? AND department = ?",
                (tender_ref, department)
            )
            count = cursor.fetchone()[0]
            conn.close()
            return count > 0
        except Exception as e:
            logger.warning(f"⚠️ Error checking tender duplicate: {e}")
            return False

    def is_file_duplicate(self, file_path, tender_ref, version=""):
        """Check if file is duplicate by hash and metadata"""
        try:
            if not os.path.exists(file_path):
                return False

            file_hash = self.calculate_file_hash(file_path)
            if not file_hash:
                return False

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute(
                "SELECT file_path FROM downloaded_files WHERE file_hash = ?",
                (file_hash,)
            )
            hash_match = cursor.fetchone()

            if hash_match:
                logger.info(f"🔄 Exact duplicate detected: {os.path.basename(file_path)}")
                conn.close()
                return True

            cursor.execute(
                "SELECT file_path FROM downloaded_files WHERE tender_ref = ? AND version = ?",
                (tender_ref, version)
            )
            ref_match = cursor.fetchone()
            conn.close()

            return ref_match is not None

        except Exception as e:
            logger.warning(f"⚠️ Error checking file duplicate: {e}")
            return False

    def record_download(self, tender_ref, file_path, department, version="", tender_data=None, ocr_processed=False):
        """Record successful download in database"""
        try:
            if not os.path.exists(file_path):
                return False

            file_hash = self.calculate_file_hash(file_path)
            file_size = os.path.getsize(file_path)

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('''
                INSERT OR REPLACE INTO downloaded_files
                (tender_ref, file_path, file_hash, file_size, department, version, ocr_processed)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (tender_ref, file_path, file_hash, file_size, department, version, 1 if ocr_processed else 0))

            if tender_data:
                cursor.execute('''
                    INSERT OR REPLACE INTO tender_metadata
                    (tender_ref, title, department, posting_date, closing_date, amount, ocr_confidence, items_count)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    tender_ref,
                    tender_data.get('title', ''),
                    department,
                    tender_data.get('posting_date', ''),
                    tender_data.get('closing_date', ''),
                    tender_data.get('amount', ''),
                    tender_data.get('ocr_confidence', 0),
                    tender_data.get('items_count', 0)
                ))

            conn.commit()
            conn.close()
            logger.info(f"✅ Recorded: {os.path.basename(file_path)} ({tender_ref})")
            return True

        except Exception as e:
            logger.error(f"❌ Error recording download: {e}")
            return False

    def get_download_stats(self):
        """Get download statistics"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute("SELECT COUNT(*) FROM downloaded_files")
            total_files = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM downloaded_files WHERE ocr_processed = 1")
            ocr_processed = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM tender_metadata")
            total_tenders = cursor.fetchone()[0]

            cursor.execute(
                "SELECT COUNT(*) FROM downloaded_files WHERE download_date > datetime('now', '-1 day')"
            )
            recent_downloads = cursor.fetchone()[0]

            conn.close()

            return {
                'total_files': total_files,
                'total_tenders': total_tenders,
                'recent_downloads': recent_downloads,
                'ocr_processed': ocr_processed
            }

        except Exception as e:
            logger.warning(f"⚠️ Error getting stats: {e}")
            return {}

# Global duplicate monitor
duplicate_monitor = None

def initialize_duplicate_monitor():
    """Initialize global duplicate monitor"""
    global duplicate_monitor
    duplicate_monitor = DuplicateMonitor(CONFIG["base_folder"])
    stats = duplicate_monitor.get_download_stats()
    if stats.get('total_files', 0) > 0:
        logger.info(f"📊 Duplicate Monitor: {stats.get('total_files', 0)} files, {stats.get('total_tenders', 0)} tenders tracked")
    return duplicate_monitor

# ============================================================================
# OCR ENGINE (Tesseract-based)
# ============================================================================

class OCREngine:
    """Handles OCR operations on PDF files using Tesseract"""

    def __init__(self, language: str = "eng+ara", dpi: int = 300):
        self.language = language
        self.dpi = dpi
        self.available = OCR_AVAILABLE
        if self.available:
            self._setup()

    def _setup(self):
        """Initialize OCR engine"""
        self.tesseract_config = r'--oem 3 --psm 6'
        # Try to find tesseract
        tesseract_paths = [
            "/opt/homebrew/bin/tesseract",
            "/usr/local/bin/tesseract",
            "/usr/bin/tesseract",
            "tesseract"
        ]
        for path in tesseract_paths:
            try:
                subprocess.run([path, "--version"], capture_output=True, check=True)
                pytesseract.pytesseract.tesseract_cmd = path
                logger.info(f"✅ Tesseract found at: {path}")
                return
            except:
                continue
        logger.warning("⚠️ Tesseract not found in standard paths")

    def preprocess_image(self, image):
        """Preprocess image for better OCR accuracy"""
        if not self.available:
            return image
        try:
            # Convert to grayscale
            if image.mode != 'L':
                image = image.convert('L')
            # Enhance contrast
            enhancer = ImageEnhance.Contrast(image)
            image = enhancer.enhance(1.5)
            # Sharpen
            image = image.filter(ImageFilter.SHARPEN)
            # Binarize
            image = image.point(lambda p: 255 if p > 128 else 0)
            return image
        except Exception as e:
            logger.warning(f"Image preprocessing failed: {e}")
            return image

    def pdf_to_images(self, pdf_path: str, max_pages: int = 10) -> List:
        """Convert PDF pages to images"""
        if not self.available:
            return []
        try:
            images = convert_from_path(
                pdf_path,
                dpi=self.dpi,
                first_page=1,
                last_page=max_pages
            )
            logger.info(f"📄 Converted {len(images)} pages from {os.path.basename(pdf_path)}")
            return images
        except Exception as e:
            logger.error(f"❌ Error converting PDF to images: {e}")
            return []

    def extract_text_from_image(self, image) -> Tuple[str, float]:
        """Extract text from a single image with confidence score"""
        if not self.available:
            return "", 0.0
        try:
            # Preprocess image
            processed = self.preprocess_image(image)

            # Get detailed OCR data including confidence
            ocr_data = pytesseract.image_to_data(
                processed,
                lang=self.language,
                config=self.tesseract_config,
                output_type=pytesseract.Output.DICT
            )

            # Calculate average confidence
            confidences = [int(c) for c in ocr_data['conf'] if int(c) > 0]
            avg_confidence = sum(confidences) / len(confidences) if confidences else 0

            # Get full text
            text = pytesseract.image_to_string(
                processed,
                lang=self.language,
                config=self.tesseract_config
            )

            return text, avg_confidence
        except Exception as e:
            logger.error(f"❌ OCR error: {e}")
            return "", 0.0

    def extract_text_from_pdf(self, pdf_path: str, max_pages: int = 10) -> Tuple[str, float]:
        """Extract all text from PDF using OCR"""
        if not self.available:
            # Fallback to PyPDF2
            return self._extract_with_pypdf2(pdf_path)

        images = self.pdf_to_images(pdf_path, max_pages)
        if not images:
            return self._extract_with_pypdf2(pdf_path)

        all_text = []
        all_confidences = []

        for i, image in enumerate(images):
            logger.info(f"  📖 OCR processing page {i+1}/{len(images)}...")
            text, confidence = self.extract_text_from_image(image)
            all_text.append(f"--- PAGE {i+1} ---\n{text}")
            all_confidences.append(confidence)

        combined_text = "\n\n".join(all_text)
        avg_confidence = sum(all_confidences) / len(all_confidences) if all_confidences else 0

        logger.info(f"✅ OCR complete. Confidence: {avg_confidence:.1f}%")
        return combined_text, avg_confidence

    def _extract_with_pypdf2(self, pdf_path: str) -> Tuple[str, float]:
        """Fallback text extraction using PyPDF2"""
        try:
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() + "\n"
                logger.info(f"📄 PyPDF2 extracted {len(text)} characters")
                return text, 50.0  # Default confidence for non-OCR
        except Exception as e:
            logger.error(f"PyPDF2 extraction failed: {e}")
            return "", 0.0

# ============================================================================
# TEXT PARSING AND EXTRACTION
# ============================================================================

def detect_text_language(text: str) -> Dict:
    """Detect if text contains Arabic content"""
    arabic_pattern = r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\uFB50-\uFDFF\uFE70-\uFEFF]'
    arabic_chars = len(re.findall(arabic_pattern, text))
    total_chars = len(re.sub(r'\s+', '', text))

    if total_chars == 0:
        return {'language': 'unknown', 'arabic_ratio': 0, 'has_arabic': False}

    arabic_ratio = arabic_chars / total_chars

    if arabic_ratio > 0.3:
        primary_language = 'arabic'
    elif arabic_ratio > 0.05:
        primary_language = 'mixed'
    else:
        primary_language = 'english'

    return {
        'language': primary_language,
        'arabic_ratio': arabic_ratio,
        'has_arabic': arabic_chars > 0
    }

def extract_reference_number(text: str, filename: str = "") -> str:
    """Extract tender reference number with improved patterns"""
    # Try filename first - common MOH format like 5AN063
    if filename:
        # MOH standard format: digit + 2-4 letters + digits
        match = re.search(r'(\d{1,2}[A-Z]{2,4}\d{2,4})', filename, re.IGNORECASE)
        if match:
            return match.group(1).upper()

    # Arabic tender reference format: رقم الممارسة or similar
    arabic_patterns = [
        r'(?:الممارسة\s*رقم|رقم\s*المناقصة|مناقصة\s*رقم)[^\d]*(\d{4}[-/]\d{4}/\d+)',
        r'(?:ه\.ط\.\s*\(?)(\d{4}[-/]\d{4}/\d+)',
    ]

    for pattern in arabic_patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)

    # English patterns for tender reference
    patterns = [
        # Standard MOH format: 1-2 digits + 2-4 letters + 2-4 digits
        r'\b(\d{1,2}[A-Z]{2,4}\d{2,4})\b',
        # Tender No/Number/Ref format
        r'(?:Tender|Bid|Practice)\s*(?:No\.?|Number|Ref\.?)?[:\s#]*(\d{1,2}[A-Z]{2,4}\d{2,4})',
        # Year-based format: 2024-2025/123
        r'(\d{4}[-/]\d{4}/\d+)',
        # Common department codes
        r'\b(\d{1,2}(?:TN|LB|AL|EQ|LS|MA|PS|PT|TE|TS|IC|RC|BM|AN|MS|BE)\d{2,4})\b',
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).upper()

    return ""

def extract_closing_date(text: str) -> str:
    """Extract closing date from tender text"""
    patterns = [
        r'(?:Closing\s*Date|Close\s*Date|Last\s*Date|Deadline)[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
        r'(?:close[sd]?\s*(?:on|by)?|deadline)[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
        r'(\d{2}/\d{2}/\d{4})',
    ]

    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            date_str = matches[-1] if len(matches) > 1 else matches[0]
            return date_str.replace('-', '/').replace('.', '/')

    return ""

def is_dimension_or_measurement(text: str, quantity: str) -> bool:
    """Check if the extracted 'quantity' is actually a dimension/measurement

    We need to be careful not to filter out real quantities. The key is:
    - If the quantity is followed directly by a unit (cm, mm, etc), it's a dimension
    - If the quantity appears after "QTY:" it's almost certainly a real quantity
    """
    # If text contains "QTY:" followed by the quantity, it's definitely a real quantity
    if re.search(r'QTY[:\s]*' + re.escape(quantity), text, re.IGNORECASE):
        return False

    # Check if quantity appears with a unit of measurement immediately after
    # Look for patterns like "162.6 cm" or "45 mm" where the number is our quantity
    quantity_with_unit = re.compile(
        re.escape(quantity) + r'\s*(?:cm|mm|m|in|inch|inches|ft|feet|°|degrees?|kg|g|lb|lbs|%)',
        re.IGNORECASE
    )
    if quantity_with_unit.search(text):
        return True

    # Check if quantity is part of dimension format like "60 x 24"
    if re.search(re.escape(quantity) + r'\s*x\s*\d+', text) or \
       re.search(r'\d+\s*x\s*' + re.escape(quantity), text):
        return True

    return False


def clean_item_description(description: str) -> str:
    """Clean and normalize item description"""
    # Remove leading numbers/bullets
    description = re.sub(r'^[\d\s.)\-:]+', '', description)
    # Remove trailing garbage
    description = re.sub(r'\s*[-–:]\s*$', '', description)
    # Normalize whitespace
    description = re.sub(r'\s+', ' ', description).strip()
    # Remove common OCR artifacts
    description = re.sub(r'[|_]{2,}', '', description)
    return description[:300]


def extract_items_from_text(text: str) -> List[TenderItem]:
    """Extract items with descriptions and quantities from OCR text - IMPROVED VERSION

    This function handles MOH tender documents with patterns like:
    - "1. ITEM NAME QTY: 2"
    - "Item 1: Description - Quantity: 10 units"
    - Tabular formats with item numbers and quantities
    """
    items = []
    seen_descriptions = set()
    seen_item_numbers = set()

    # PRIMARY PATTERN: MOH standard format "1. ITEM DESCRIPTION QTY: N" or "QTY: N"
    # This is the most reliable pattern for MOH tenders
    qty_pattern = re.compile(
        r'(\d{1,3})\s*[.)]\s*([A-Z][^\n]{10,200}?)\s+(?:QTY|Qty|QUANTITY|Quantity)[:\s]*(\d+)',
        re.IGNORECASE
    )

    for match in qty_pattern.finditer(text):
        item_no = match.group(1)
        description = clean_item_description(match.group(2))
        quantity = match.group(3)

        # Skip if this looks like a dimension
        if is_dimension_or_measurement(match.group(0), quantity):
            continue

        # Skip duplicates by item number
        if item_no in seen_item_numbers:
            continue

        desc_key = description.lower()[:80]
        if desc_key in seen_descriptions or len(description) < 10:
            continue

        seen_item_numbers.add(item_no)
        seen_descriptions.add(desc_key)
        lang_info = detect_text_language(description)

        items.append(TenderItem(
            item_number=item_no,
            description=description,
            quantity=quantity,
            unit="units",
            language=lang_info['language'],
            has_arabic=lang_info['has_arabic']
        ))

    # SECONDARY PATTERN: Items with explicit quantity units
    if len(items) < 3:
        unit_pattern = re.compile(
            r'(\d{1,3})\s*[.)]\s*([A-Za-z][^\n]{10,150}?)\s+(\d+)\s*(pieces?|pcs?|units?|sets?|nos?|each)',
            re.IGNORECASE
        )

        for match in unit_pattern.finditer(text):
            item_no = match.group(1)
            description = clean_item_description(match.group(2))
            quantity = match.group(3)
            unit = match.group(4).lower()

            if is_dimension_or_measurement(match.group(0), quantity):
                continue

            if item_no in seen_item_numbers:
                continue

            desc_key = description.lower()[:80]
            if desc_key in seen_descriptions or len(description) < 10:
                continue

            seen_item_numbers.add(item_no)
            seen_descriptions.add(desc_key)
            lang_info = detect_text_language(description)

            items.append(TenderItem(
                item_number=item_no,
                description=description,
                quantity=quantity,
                unit=unit,
                language=lang_info['language'],
                has_arabic=lang_info['has_arabic']
            ))

    # TERTIARY PATTERN: Arabic items with عدد (quantity)
    arabic_pattern = re.compile(
        r'(\d{1,3})\s*[.)]\s*([\u0600-\u06FF][^\n]{10,150}?)\s+(?:عدد|الكمية)[:\s]*(\d+)',
        re.IGNORECASE
    )

    for match in arabic_pattern.finditer(text):
        item_no = match.group(1)
        description = clean_item_description(match.group(2))
        quantity = match.group(3)

        if item_no in seen_item_numbers:
            continue

        desc_key = description.lower()[:80]
        if desc_key in seen_descriptions or len(description) < 10:
            continue

        seen_item_numbers.add(item_no)
        seen_descriptions.add(desc_key)

        items.append(TenderItem(
            item_number=item_no,
            description=description,
            quantity=quantity,
            unit="وحدة",
            language='arabic',
            has_arabic=True
        ))

    # QUATERNARY PATTERN: Medical Store tabular format "1 ITEM_NAME UNIT QUANTITY"
    # Example: "1 EPHEDRINE 3MG/ML PFS 30,000"
    # Also handles: "1 | FASTTHREAD INTERFERENCE SCREW PCS 50"
    # Common units: PFS, VIAL, AMP, TAB, CAP, BOT, PCS, KIT, etc.
    if len(items) < 3:
        tabular_pattern = re.compile(
            r'^(\d{1,3})\s*\|?\s*([A-Z][A-Z0-9\s\-/.,()%#]+?)\s+(PFS|VIAL|AMP|TAB|CAP|BOT|BTL|PKT|TUBE|BAG|PCK|SET|EACH|EA|UNIT|BOX|PACK|PCS|KIT|PAIR|ROLL)\s+([\d,]+)$',
            re.IGNORECASE | re.MULTILINE
        )

        for match in tabular_pattern.finditer(text):
            item_no = match.group(1)
            description = match.group(2).strip()
            unit = match.group(3).upper()
            quantity = match.group(4).replace(',', '')

            if item_no in seen_item_numbers:
                continue

            # Clean up description
            description = re.sub(r'\s+', ' ', description).strip()
            if len(description) < 5:
                continue

            desc_key = description.lower()[:80]
            if desc_key in seen_descriptions:
                continue

            seen_item_numbers.add(item_no)
            seen_descriptions.add(desc_key)
            lang_info = detect_text_language(description)

            items.append(TenderItem(
                item_number=item_no,
                description=description,
                quantity=quantity,
                unit=unit,
                language=lang_info['language'],
                has_arabic=lang_info['has_arabic']
            ))

    # FALLBACK: Line-by-line extraction for simpler formats
    if len(items) < 2:
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if len(line) < 15 or len(line) > 200:
                continue

            # Look for "Item Name - Qty: N" or similar
            simple_match = re.search(
                r'^(\d{1,2})\s*[.)]\s*(.{15,100}?)\s*[-–]\s*(\d+)\s*(?:units?|pcs?)?$',
                line, re.IGNORECASE
            )

            if simple_match:
                item_no = simple_match.group(1)
                description = clean_item_description(simple_match.group(2))
                quantity = simple_match.group(3)

                if item_no in seen_item_numbers:
                    continue

                desc_key = description.lower()[:80]
                if desc_key in seen_descriptions or len(description) < 10:
                    continue

                if is_dimension_or_measurement(line, quantity):
                    continue

                seen_item_numbers.add(item_no)
                seen_descriptions.add(desc_key)
                lang_info = detect_text_language(description)

                items.append(TenderItem(
                    item_number=item_no,
                    description=description,
                    quantity=quantity,
                    unit="units",
                    language=lang_info['language'],
                    has_arabic=lang_info['has_arabic']
                ))

    # Sort by item number
    try:
        items.sort(key=lambda x: int(x.item_number) if x.item_number.isdigit() else 999)
    except:
        pass

    logger.info(f"🎯 Extracted {len(items)} unique items from text")
    return items[:100]

def extract_specifications(text: str) -> str:
    """Extract specifications section"""
    patterns = [
        r'(?:Technical\s*)?Specifications?[:\s]*(.+?)(?:Terms|Conditions|Notes|\Z)',
        r'Requirements?[:\s]*(.+?)(?:Terms|Notes|\Z)',
        r'المواصفات[:\s]*(.+?)(?:الشروط|\Z)',
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            specs = match.group(1).strip()
            specs = re.sub(r'\s+', ' ', specs)
            if len(specs) > 50:
                return specs[:2000]

    return ""

# ============================================================================
# SELENIUM HELPER FUNCTIONS
# ============================================================================

def human_delay(min_time=1.0, max_time=3.0, fast_mode=None):
    """Add random human-like delay"""
    if fast_mode is None:
        fast_mode = CONFIG.get("fast_mode", True)
    if fast_mode:
        time.sleep(random.uniform(0.3, 0.8))
    else:
        time.sleep(random.uniform(min_time, max_time))

def human_scroll(driver, scroll_amount=None):
    """Perform human-like scrolling"""
    try:
        if scroll_amount is None:
            scroll_amount = random.randint(200, 400)
        driver.execute_script(f"window.scrollBy(0, {scroll_amount})")
        human_delay(0.2, 0.5)
    except:
        pass

def setup_driver(headless=None):
    """Setup Chrome WebDriver with anti-detection measures"""
    if headless is None:
        headless = CONFIG.get("headless_mode", True)

    logger.info("🚀 Setting up Chrome WebDriver...")
    chromedriver_autoinstaller.install()

    options = webdriver.ChromeOptions()

    if headless:
        options.add_argument('--headless=new')

    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')

    # Disable automation flags
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    # Set download preferences
    download_dir = CONFIG.get("output_folder", os.path.expanduser("~/Desktop/MOH Tenders/scraping_results"))
    os.makedirs(download_dir, exist_ok=True)

    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True,
        "safebrowsing.enabled": True
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=options)
    driver.execute_cdp_cmd('Network.setUserAgentOverride', {
        "userAgent": 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    })

    # Hide webdriver property
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    logger.info("✅ WebDriver ready")
    return driver

def safe_click(driver, element, wait_time=0.5):
    """Safely click an element with retry logic"""
    try:
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        time.sleep(wait_time)
        driver.execute_script("arguments[0].click();", element)
        return True
    except Exception as e:
        try:
            ActionChains(driver).move_to_element(element).click().perform()
            return True
        except Exception as e2:
            logger.warning(f"⚠️ Click failed: {e2}")
            return False

def wait_for_element(driver, selector, by=By.CSS_SELECTOR, timeout=10):
    """Wait for element to be visible"""
    try:
        wait = WebDriverWait(driver, timeout)
        element = wait.until(EC.visibility_of_element_located((by, selector)))
        return element
    except Exception as e:
        logger.warning(f"⚠️ Element not found: {selector}")
        return None

# ============================================================================
# TENDER SCRAPING FUNCTIONS
# ============================================================================

def navigate_to_tenders(driver, department):
    """Navigate to the tenders page for a specific department"""
    try:
        home_url = CONFIG.get("home_url", "https://www.moh.gov.kw/en/Pages/default.aspx")
        logger.info(f"🌐 Navigating to MOH website for {department}...")
        driver.get(home_url)
        human_delay(2, 4)

        # Look for tenders/bids links
        tender_keywords = ['tender', 'bid', 'procurement', 'مناقصة', 'عطاء']
        links = driver.find_elements(By.TAG_NAME, 'a')

        for link in links:
            href = link.get_attribute('href') or ''
            text = link.text.lower()

            for keyword in tender_keywords:
                if keyword in href.lower() or keyword in text:
                    logger.info(f"🔗 Found tender link: {link.text[:50]}")
                    safe_click(driver, link)
                    human_delay(2, 3)
                    return True

        # Try department-specific navigation
        dept_map = {
            'Medical Store': ['medical', 'store', 'pharmacy'],
            'Biomedical Engineering': ['biomedical', 'engineering', 'equipment']
        }

        for key_dept, keywords in dept_map.items():
            if department.lower() in key_dept.lower():
                for link in links:
                    text = link.text.lower()
                    if any(kw in text for kw in keywords):
                        safe_click(driver, link)
                        human_delay(1, 2)
                        break

        return True

    except Exception as e:
        logger.error(f"❌ Navigation error: {e}")
        return False

def extract_tender_links(driver):
    """Extract all tender links from current page"""
    tender_links = []

    try:
        soup = BeautifulSoup(driver.page_source, 'html.parser')

        # Find PDF links
        pdf_links = soup.find_all('a', href=re.compile(r'\.pdf', re.IGNORECASE))
        for link in pdf_links:
            href = link.get('href', '')
            if href:
                full_url = urljoin(driver.current_url, href)
                text = link.get_text(strip=True) or os.path.basename(href)
                tender_links.append({
                    'url': full_url,
                    'title': text,
                    'type': 'pdf'
                })

        # Find tender detail pages
        detail_patterns = [
            r'tender.*detail', r'view.*tender', r'tender.*info',
            r'bid.*detail', r'\?id=\d+', r'itemid=\d+'
        ]

        all_links = soup.find_all('a', href=True)
        for link in all_links:
            href = link.get('href', '')
            for pattern in detail_patterns:
                if re.search(pattern, href, re.IGNORECASE):
                    full_url = urljoin(driver.current_url, href)
                    text = link.get_text(strip=True)
                    tender_links.append({
                        'url': full_url,
                        'title': text,
                        'type': 'page'
                    })
                    break

        logger.info(f"🔗 Found {len(tender_links)} tender links")
        return tender_links

    except Exception as e:
        logger.error(f"❌ Error extracting links: {e}")
        return []

def download_pdf(url, save_path, timeout=60):
    """Download PDF file with progress tracking"""
    try:
        logger.info(f"📥 Downloading: {os.path.basename(save_path)}")

        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'application/pdf,*/*',
        }

        response = requests.get(url, headers=headers, stream=True, timeout=timeout)
        response.raise_for_status()

        content_type = response.headers.get('Content-Type', '')
        if 'pdf' not in content_type.lower() and not url.lower().endswith('.pdf'):
            logger.warning(f"⚠️ Not a PDF: {content_type}")
            return False

        os.makedirs(os.path.dirname(save_path), exist_ok=True)

        with open(save_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

        file_size = os.path.getsize(save_path)
        if file_size < 1000:
            logger.warning(f"⚠️ File too small ({file_size} bytes), may be invalid")
            os.remove(save_path)
            return False

        logger.info(f"✅ Downloaded: {os.path.basename(save_path)} ({file_size/1024:.1f} KB)")
        return True

    except Exception as e:
        logger.error(f"❌ Download failed: {e}")
        if os.path.exists(save_path):
            os.remove(save_path)
        return False

def extract_tender_ref_from_url(url):
    """Extract tender reference from URL"""
    patterns = [
        r'(\d{1,2}[A-Z]{2,3}\d{2,4})',
        r'tender[_-]?(\d+)',
        r'id[_=](\d+)',
    ]

    for pattern in patterns:
        match = re.search(pattern, url, re.IGNORECASE)
        if match:
            return match.group(1)

    # Use filename as fallback
    parsed = urlparse(url)
    filename = os.path.basename(parsed.path)
    name = os.path.splitext(filename)[0]
    return re.sub(r'[^\w\-]', '_', name)[:50]

def scrape_department_tenders(driver, department, max_tenders=50):
    """Scrape all tenders for a department"""
    global duplicate_monitor

    results = []
    ocr_engine = OCREngine(CONFIG.get("tesseract_lang", "eng+ara"), CONFIG.get("dpi", 300))

    logger.info(f"\n{'='*60}")
    logger.info(f"📋 Scraping department: {department}")
    logger.info(f"{'='*60}")

    # Navigate to tenders
    if not navigate_to_tenders(driver, department):
        logger.error(f"❌ Failed to navigate to {department}")
        return results

    human_delay(2, 4)

    # Extract tender links
    tender_links = extract_tender_links(driver)

    if not tender_links:
        logger.warning(f"⚠️ No tender links found for {department}")
        return results

    # Process each tender
    processed = 0
    for tender_info in tender_links[:max_tenders]:
        try:
            url = tender_info['url']
            title = tender_info['title']
            tender_ref = extract_tender_ref_from_url(url)

            logger.info(f"\n📄 Processing: {tender_ref}")

            # Check for duplicates
            if duplicate_monitor and duplicate_monitor.is_tender_duplicate(tender_ref, department):
                logger.info(f"⏭️ Skipping duplicate: {tender_ref}")
                continue

            # Download PDF - use department-specific folders matching existing structure
            if "biomedical" in department.lower():
                output_dir = CONFIG["biomedical_folder"]
            elif "medical" in department.lower() and "store" in department.lower():
                output_dir = CONFIG["medical_store_folder"]
            else:
                output_dir = os.path.join(CONFIG["base_folder"], department.replace(" ", "_"))
            os.makedirs(output_dir, exist_ok=True)

            if tender_info['type'] == 'pdf':
                pdf_path = os.path.join(output_dir, f"{tender_ref}.pdf")

                if os.path.exists(pdf_path):
                    logger.info(f"📁 File exists: {pdf_path}")
                else:
                    if not download_pdf(url, pdf_path):
                        continue

                # Run OCR if enabled
                tender_data = TenderData(
                    reference=tender_ref,
                    department=department,
                    title=title,
                    source_files=[pdf_path]
                )

                if CONFIG.get("enable_ocr", True) and ocr_engine.available:
                    logger.info(f"🔍 Running OCR on {tender_ref}...")
                    raw_text, confidence = ocr_engine.extract_text_from_pdf(
                        pdf_path,
                        CONFIG.get("max_pages", 10)
                    )

                    tender_data.raw_text = raw_text
                    tender_data.ocr_confidence = confidence

                    # Extract structured data
                    tender_data.reference = extract_reference_number(raw_text, tender_ref) or tender_ref
                    tender_data.closing_date = extract_closing_date(raw_text)
                    tender_data.items = extract_items_from_text(raw_text)
                    tender_data.specifications = extract_specifications(raw_text)

                    lang_info = detect_text_language(raw_text)
                    tender_data.has_arabic_content = lang_info['has_arabic']

                    logger.info(f"✅ OCR complete: {len(tender_data.items)} items, {confidence:.1f}% confidence")

                # Record in database
                if duplicate_monitor:
                    duplicate_monitor.record_download(
                        tender_ref, pdf_path, department, "",
                        {
                            'title': title,
                            'closing_date': tender_data.closing_date,
                            'ocr_confidence': tender_data.ocr_confidence,
                            'items_count': len(tender_data.items)
                        },
                        ocr_processed=CONFIG.get("enable_ocr", True)
                    )

                results.append(tender_data)
                processed += 1

            else:
                # Handle tender detail pages
                driver.get(url)
                human_delay(1, 2)

                # Look for PDF links on detail page
                page_pdf_links = extract_tender_links(driver)
                for pdf_info in page_pdf_links:
                    if pdf_info['type'] == 'pdf':
                        pdf_ref = extract_tender_ref_from_url(pdf_info['url'])
                        pdf_path = os.path.join(output_dir, f"{pdf_ref}.pdf")

                        if download_pdf(pdf_info['url'], pdf_path):
                            tender_data = TenderData(
                                reference=pdf_ref,
                                department=department,
                                title=pdf_info['title'],
                                source_files=[pdf_path]
                            )

                            if CONFIG.get("enable_ocr", True) and ocr_engine.available:
                                raw_text, confidence = ocr_engine.extract_text_from_pdf(pdf_path, CONFIG.get("max_pages", 10))
                                tender_data.raw_text = raw_text
                                tender_data.ocr_confidence = confidence
                                tender_data.items = extract_items_from_text(raw_text)
                                tender_data.closing_date = extract_closing_date(raw_text)

                            results.append(tender_data)
                            processed += 1

            human_delay(1, 2)

        except Exception as e:
            logger.error(f"❌ Error processing tender: {e}")
            continue

    logger.info(f"\n✅ {department}: Processed {processed} tenders")
    return results

# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def export_to_json(tenders: List[TenderData], output_path: str):
    """Export tender data to JSON"""
    try:
        data = []
        for tender in tenders:
            tender_dict = asdict(tender)
            tender_dict['items'] = [asdict(item) for item in tender.items]
            data.append(tender_dict)

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        logger.info(f"📄 Exported JSON: {output_path}")
        return True
    except Exception as e:
        logger.error(f"❌ JSON export failed: {e}")
        return False

def export_to_excel(tenders: List[TenderData], output_path: str):
    """Export all tender data to a single comprehensive Excel sheet"""
    try:
        # Create one row per item with all tender info included
        all_data = []

        for tender in tenders:
            # Base tender info that repeats for each item
            base_info = {
                'Tender Reference': tender.reference,
                'Department': tender.department,
                'Title': tender.title,
                'Closing Date': tender.closing_date,
                'Total Items': len(tender.items),
                'OCR Confidence': f"{tender.ocr_confidence:.1f}%",
                'Has Arabic': 'Yes' if tender.has_arabic_content else 'No',
                'Extraction Date': tender.extraction_date,
                'Specifications': tender.specifications[:500] if tender.specifications else '',
                'Source File': tender.source_files[0] if tender.source_files else ''
            }

            if tender.items:
                # Create a row for each item
                for item in tender.items:
                    row = base_info.copy()
                    row.update({
                        'Item #': item.item_number,
                        'Item Description': item.description,
                        'Quantity': item.quantity,
                        'Unit': item.unit,
                        'Item Language': item.language,
                        'Item Has Arabic': 'Yes' if item.has_arabic else 'No'
                    })
                    all_data.append(row)
            else:
                # No items - still add tender info with empty item fields
                row = base_info.copy()
                row.update({
                    'Item #': '',
                    'Item Description': 'No items extracted',
                    'Quantity': '',
                    'Unit': '',
                    'Item Language': '',
                    'Item Has Arabic': ''
                })
                all_data.append(row)

        # Create Excel workbook with single sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Tender Data"

        if all_data:
            # Define column order for better readability
            column_order = [
                'Tender Reference', 'Department', 'Title', 'Closing Date',
                'Item #', 'Item Description', 'Quantity', 'Unit',
                'Total Items', 'OCR Confidence', 'Has Arabic', 'Item Has Arabic',
                'Specifications', 'Extraction Date', 'Source File'
            ]

            # Ensure all columns exist
            for row in all_data:
                for col in column_order:
                    if col not in row:
                        row[col] = ''

            df = pd.DataFrame(all_data, columns=column_order)

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        # Header styling
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    else:
                        # Alternate row colors for readability
                        if r_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

            # Adjust column widths
            column_widths = {
                'Tender Reference': 20,
                'Department': 20,
                'Title': 40,
                'Closing Date': 15,
                'Item #': 8,
                'Item Description': 50,
                'Quantity': 12,
                'Unit': 12,
                'Total Items': 12,
                'OCR Confidence': 15,
                'Has Arabic': 12,
                'Item Has Arabic': 15,
                'Specifications': 60,
                'Extraction Date': 20,
                'Source File': 40
            }

            for c_idx, col_name in enumerate(column_order, 1):
                col_letter = openpyxl.utils.get_column_letter(c_idx)
                ws.column_dimensions[col_letter].width = column_widths.get(col_name, 15)

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Add auto-filter
            ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)
        logger.info(f"📊 Exported Excel (single sheet): {output_path} - {len(all_data)} rows")
        return True

    except Exception as e:
        logger.error(f"❌ Excel export failed: {e}")
        return False

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def load_progress_tracker(progress_file: str) -> dict:
    """Load progress tracker from file"""
    if os.path.exists(progress_file):
        try:
            with open(progress_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Could not load progress file: {e}")
    return {"processed_files": [], "failed_files": [], "total_processed": 0}


def save_progress_tracker(progress_file: str, tracker: dict):
    """Save progress tracker to file"""
    try:
        with open(progress_file, 'w') as f:
            json.dump(tracker, f, indent=2)
    except Exception as e:
        logger.error(f"Could not save progress file: {e}")


def format_time(seconds: float) -> str:
    """Format seconds into human-readable time"""
    if seconds < 60:
        return f"{seconds:.0f}s"
    elif seconds < 3600:
        mins = seconds // 60
        secs = seconds % 60
        return f"{mins:.0f}m {secs:.0f}s"
    else:
        hours = seconds // 3600
        mins = (seconds % 3600) // 60
        return f"{hours:.0f}h {mins:.0f}m"


def process_existing_pdfs(pdf_folder: str, output_folder: str = None, batch_size: int = 50):
    """Process existing PDFs with OCR (batch mode with progress tracking and incremental saves)

    Features:
    - Batch processing with incremental exports every batch_size files
    - Progress tracking to skip already-processed PDFs (resume capability)
    - ETA calculation and progress monitoring
    - Periodic checkpoint saves to prevent data loss
    """
    if output_folder is None:
        output_folder = CONFIG.get("ocr_output_folder", os.path.join(pdf_folder, "ocr_results"))

    os.makedirs(output_folder, exist_ok=True)

    # Progress tracking file
    progress_file = os.path.join(output_folder, "ocr_progress.json")
    tracker = load_progress_tracker(progress_file)
    processed_set = set(tracker.get("processed_files", []))

    ocr_engine = OCREngine(CONFIG.get("tesseract_lang", "eng+ara"), CONFIG.get("dpi", 300))
    results = []
    batch_results = []

    # Find all PDFs
    pdf_files = []
    for root, dirs, files in os.walk(pdf_folder):
        for file in files:
            if file.lower().endswith('.pdf'):
                pdf_files.append(os.path.join(root, file))

    total_files = len(pdf_files)

    # Filter out already processed files
    pending_files = [f for f in pdf_files if f not in processed_set]
    skipped_count = total_files - len(pending_files)

    logger.info(f"\n{'='*60}")
    logger.info(f"📁 BATCH OCR PROCESSING")
    logger.info(f"{'='*60}")
    logger.info(f"   Total PDF files found: {total_files}")
    logger.info(f"   Already processed (skipping): {skipped_count}")
    logger.info(f"   Remaining to process: {len(pending_files)}")
    logger.info(f"   Batch size: {batch_size} files")
    logger.info(f"   Progress file: {progress_file}")
    logger.info(f"{'='*60}\n")

    if not pending_files:
        logger.info("✅ All files already processed! Nothing to do.")
        return results

    # Timing for ETA
    start_time = time.time()
    processing_times = []
    batch_num = 0
    session_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    for i, pdf_path in enumerate(pending_files):
        file_start_time = time.time()

        try:
            # Progress with ETA
            if processing_times:
                avg_time = sum(processing_times) / len(processing_times)
                remaining = len(pending_files) - i
                eta_seconds = avg_time * remaining
                eta_str = format_time(eta_seconds)
                elapsed = time.time() - start_time
                elapsed_str = format_time(elapsed)
                progress_pct = ((i + skipped_count) / total_files) * 100

                logger.info(f"\n[{i+1}/{len(pending_files)}] ({progress_pct:.1f}% overall) ETA: {eta_str} | Elapsed: {elapsed_str}")
            else:
                logger.info(f"\n[{i+1}/{len(pending_files)}] Processing...")

            logger.info(f"   📄 {os.path.basename(pdf_path)}")

            tender_ref = extract_tender_ref_from_url(pdf_path)

            tender_data = TenderData(
                reference=tender_ref,
                source_files=[pdf_path]
            )

            if ocr_engine.available:
                raw_text, confidence = ocr_engine.extract_text_from_pdf(pdf_path, CONFIG.get("max_pages", 10))

                tender_data.raw_text = raw_text
                tender_data.ocr_confidence = confidence
                tender_data.reference = extract_reference_number(raw_text, tender_ref) or tender_ref
                tender_data.closing_date = extract_closing_date(raw_text)
                tender_data.items = extract_items_from_text(raw_text)
                tender_data.specifications = extract_specifications(raw_text)

                lang_info = detect_text_language(raw_text)
                tender_data.has_arabic_content = lang_info['has_arabic']

                logger.info(f"   ✅ Extracted {len(tender_data.items)} items | Confidence: {confidence:.1f}%")

            results.append(tender_data)
            batch_results.append(tender_data)

            # Track as processed
            tracker["processed_files"].append(pdf_path)
            tracker["total_processed"] = len(tracker["processed_files"])

            # Record processing time
            file_time = time.time() - file_start_time
            processing_times.append(file_time)

            # Keep only last 20 times for accurate ETA
            if len(processing_times) > 20:
                processing_times.pop(0)

        except Exception as e:
            logger.error(f"   ❌ Error: {e}")
            tracker["failed_files"].append({"file": pdf_path, "error": str(e)})
            continue

        # Batch checkpoint - save every batch_size files
        if len(batch_results) >= batch_size:
            batch_num += 1
            logger.info(f"\n{'='*50}")
            logger.info(f"💾 SAVING BATCH {batch_num} ({len(batch_results)} files)")
            logger.info(f"{'='*50}")

            # Save batch results
            batch_json = os.path.join(output_folder, f"ocr_batch_{session_timestamp}_{batch_num:03d}.json")
            batch_excel = os.path.join(output_folder, f"ocr_batch_{session_timestamp}_{batch_num:03d}.xlsx")

            export_to_json(batch_results, batch_json)
            export_to_excel(batch_results, batch_excel)

            logger.info(f"   📊 Saved: {batch_excel}")
            logger.info(f"   📋 Saved: {batch_json}")

            # Save progress tracker
            save_progress_tracker(progress_file, tracker)
            logger.info(f"   📝 Progress saved ({tracker['total_processed']} total processed)")

            # Clear batch buffer (keep in full results)
            batch_results = []

    # Final save for remaining files
    if batch_results:
        batch_num += 1
        logger.info(f"\n{'='*50}")
        logger.info(f"💾 SAVING FINAL BATCH {batch_num} ({len(batch_results)} files)")
        logger.info(f"{'='*50}")

        batch_json = os.path.join(output_folder, f"ocr_batch_{session_timestamp}_{batch_num:03d}.json")
        batch_excel = os.path.join(output_folder, f"ocr_batch_{session_timestamp}_{batch_num:03d}.xlsx")

        export_to_json(batch_results, batch_json)
        export_to_excel(batch_results, batch_excel)

        logger.info(f"   📊 Saved: {batch_excel}")
        logger.info(f"   📋 Saved: {batch_json}")

    # Save final progress
    save_progress_tracker(progress_file, tracker)

    # Create combined results file
    if results:
        combined_json = os.path.join(output_folder, f"ocr_combined_{session_timestamp}.json")
        combined_excel = os.path.join(output_folder, f"ocr_combined_{session_timestamp}.xlsx")

        export_to_json(results, combined_json)
        export_to_excel(results, combined_excel)

        total_time = time.time() - start_time
        avg_time = total_time / len(results) if results else 0

        logger.info(f"\n{'='*60}")
        logger.info(f"✅ BATCH OCR COMPLETE!")
        logger.info(f"{'='*60}")
        logger.info(f"   Session processed: {len(results)} PDFs")
        logger.info(f"   Total ever processed: {tracker['total_processed']}")
        logger.info(f"   Failed: {len(tracker.get('failed_files', []))}")
        logger.info(f"   Total time: {format_time(total_time)}")
        logger.info(f"   Average per file: {format_time(avg_time)}")
        logger.info(f"   ")
        logger.info(f"   📊 Combined Excel: {combined_excel}")
        logger.info(f"   📋 Combined JSON: {combined_json}")
        logger.info(f"   📁 Batch files: {batch_num} batches saved")
        logger.info(f"{'='*60}")

    return results

def main():
    """Main entry point for the scraper"""
    global duplicate_monitor

    logger.info("\n" + "="*70)
    logger.info("🚀 MOH TENDER SCRAPER WITH INTEGRATED OCR")
    logger.info("="*70)

    # Initialize duplicate monitor
    duplicate_monitor = initialize_duplicate_monitor()

    # Create output directories
    os.makedirs(CONFIG["output_folder"], exist_ok=True)
    os.makedirs(CONFIG["ocr_output_folder"], exist_ok=True)

    all_results = []
    driver = None

    try:
        # Departments to scrape
        departments = ['Medical Store', 'Biomedical Engineering']

        # Setup WebDriver
        driver = setup_driver(CONFIG.get("headless_mode", True))

        # Scrape each department
        for department in departments:
            try:
                results = scrape_department_tenders(driver, department, max_tenders=50)
                all_results.extend(results)
            except Exception as e:
                logger.error(f"❌ Error scraping {department}: {e}")
                continue

        # Export all results
        if all_results:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            json_path = os.path.join(CONFIG["output_folder"], f"tenders_{timestamp}.json")
            excel_path = os.path.join(CONFIG["output_folder"], f"tenders_{timestamp}.xlsx")

            export_to_json(all_results, json_path)
            export_to_excel(all_results, excel_path)

        # Print summary
        logger.info("\n" + "="*70)
        logger.info("📊 SCRAPING SUMMARY")
        logger.info("="*70)
        logger.info(f"Total tenders processed: {len(all_results)}")

        stats = duplicate_monitor.get_download_stats() if duplicate_monitor else {}
        logger.info(f"Total files tracked: {stats.get('total_files', 0)}")
        logger.info(f"OCR processed: {stats.get('ocr_processed', 0)}")

        items_count = sum(len(t.items) for t in all_results)
        logger.info(f"Total items extracted: {items_count}")

        if all_results:
            avg_confidence = sum(t.ocr_confidence for t in all_results) / len(all_results)
            logger.info(f"Average OCR confidence: {avg_confidence:.1f}%")

        logger.info("="*70)

    except Exception as e:
        logger.error(f"❌ Fatal error: {e}")
        import traceback
        traceback.print_exc()

    finally:
        if driver:
            driver.quit()
            logger.info("🛑 Browser closed")

    return all_results


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="MOH Tender Scraper with Integrated OCR")
    parser.add_argument('--batch', '-b', type=str, help="Process existing PDFs in folder (batch OCR mode)")
    parser.add_argument('--output', '-o', type=str, help="Output folder for results")
    parser.add_argument('--headless', action='store_true', default=True, help="Run browser in headless mode")
    parser.add_argument('--no-headless', dest='headless', action='store_false', help="Show browser window")
    parser.add_argument('--no-ocr', action='store_true', help="Disable OCR processing")
    parser.add_argument('--dpi', type=int, default=300, help="OCR DPI setting (default: 300)")
    parser.add_argument('--pages', type=int, default=10, help="Max pages to OCR per PDF (default: 10)")

    args = parser.parse_args()

    # Update config from args
    CONFIG["headless_mode"] = args.headless
    CONFIG["enable_ocr"] = not args.no_ocr
    CONFIG["dpi"] = args.dpi
    CONFIG["max_pages"] = args.pages

    if args.output:
        CONFIG["output_folder"] = args.output
        CONFIG["ocr_output_folder"] = args.output

    if args.batch:
        # Batch OCR mode - process existing PDFs
        process_existing_pdfs(args.batch, args.output)
    else:
        # Normal scraping mode
        main()
